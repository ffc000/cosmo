"""
blueprints/vua.py — Módulo VUA (Ventanilla Única Aduanera): minutas, ejes de
trabajo, glosario, riesgos, equipo, informes normativos, BPMN, resumen
ejecutivo con IA.

Tercer blueprint extraído de app.py en la Fase 2 de profesionalización.
Sigue senasa, después finanzas.

Nota: las rutas /api/integrantes/* NO están acá aunque viven físicamente al
lado de VUA en el app.py original — las usa también SENASA (tabla
'integrantes' compartida), así que se quedan en app.py hasta que se
extraiga senasa_bp, momento en el que probablemente merezcan su propio
mini-blueprint compartido.
"""
import os
import re
import io
import json
import uuid
import time
import logging
import tempfile
import threading
import subprocess

from actas import generar_acta_word
from datetime import datetime, date, timedelta

from flask import (Blueprint, request, jsonify, render_template, session,
                    redirect, url_for, send_file)

from core import (
    HIST_DB, OUTPUT_FOLDER, login_required, modulo_required, admin_required,
    get_api_key, contexto_repositorio, notificar_telegram,
    job_status, job_create, job_get, _job_persist,
    _normalizar_fecha_a_ddmmaaaa, _validar_fecha_ddmmaaaa,
    validar_enum, ESTADOS_TAREA, NIVELES_PROBABILIDAD, NIVELES_IMPACTO, get_db,
)

vua_bp = Blueprint("vua", __name__)

# ── VUA ────────────────────────────────────────────────────────────────────────
# ROLES_PREDEFINIDOS — mantenido por compatibilidad, los datos ahora viven en tabla 'integrantes'
def get_roles_predefinidos():
    """Lee los integrantes activos de la BD y devuelve {nombre: cargo (organismo)}."""
    try:
        with get_db(HIST_DB, row_factory=True) as con:
            rows = con.execute("SELECT nombre, cargo, organismo FROM integrantes WHERE activo=1").fetchall()
        return {r["nombre"]: f"{r['cargo']} ({r['organismo']})" if r["organismo"] else r["cargo"] for r in rows}
    except Exception:
        return {}

ROLES_PREDEFINIDOS = {
    "Diego Bugallo": "Jefe Dpto. Facilitación y Simplificación de Comercio (DI REPA)",
    "Martín Macías": "Jefe Div. Modernización de Procesos Aduaneros (DI REPA)",
    "Hernán Cascón": "Supervisor de Informática Aduanera (DI SADU)",
    "Maximiliano Luengo": "Consejero técnico (DI ADEZ)",
    "Pablo Gómez Valdez": "Consejero técnico (DI ADEZ)",
    "Fabiola Cochello": "Directora VUCEA",
    "Vanesa Franco": "Jefa de Procesos VUCEA",
    "Federico Cáceres": "Sec. Simplificación de Procesos Operativos (DI REPA)",
}

CAMPOS_XFWB = [
    {"tab":"Información general","campo":"Número de guía aérea","campo_xml":"masterDocumentNumber","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":"Identificador único de la guía aérea master."},
    {"tab":"Información general","campo":"CUIT del consignatario","campo_xml":"OCI/AR/IMP//CUIT...","norma":"Art. 3° RG 4517/2019","obligatorio":True,"observacion":"Formato exacto: OCI/AR/IMP//CUIT12345678901. Sin este campo el XFWB es inválido para Aduana."},
    {"tab":"Información general","campo":"Notificación (notifyParty)","campo_xml":"ConsignmentType/NotifyParty","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":"Nombre, número de cuenta y dirección."},
    {"tab":"Información general","campo":"Agente de aduana","campo_xml":"FreightForwarder rol CustomsBroker","norma":"IATA Cargo-XML, RG 3596/2014","obligatorio":True,"observacion":"Distinto del FreightForwarderParty."},
    {"tab":"Información general","campo":"Remitente — Nombre","campo_xml":"ShipperParty/Name","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":""},
    {"tab":"Información general","campo":"Destinatario — Nombre","campo_xml":"ConsigneeParty/Name","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":""},
    {"tab":"Carga","campo":"Descripción de la mercancía","campo_xml":"GoodsDescription","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":""},
    {"tab":"Carga","campo":"Peso bruto total","campo_xml":"TotalGrossWeight","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":"En kilogramos."},
    {"tab":"Vuelo","campo":"Número de vuelo","campo_xml":"FlightBooking/FlightIdentifier","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":""},
    {"tab":"Vuelo","campo":"Aeropuerto de origen","campo_xml":"DepartureLocation","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":"Código IATA de 3 letras."},
    {"tab":"Vuelo","campo":"Aeropuerto de destino","campo_xml":"ArrivalLocation","norma":"IATA Cargo-XML XFWB schema","obligatorio":True,"observacion":"Código IATA de 3 letras."},
]

REGLAS_BPMN = {
    "EXPO": [
        {"id":"EXPO-001","descripcion":"Generación y Registración del MANE deben ser un único nodo en carril ATA MT","patron":["Generación del MANE","Registración del MANE"],"tipo":"error","norma":"RG 5756/2025"},
        {"id":"EXPO-002","descripcion":"Confirmación de Partida debe estar en carril ATA MT","patron":["Confirmación del MANE"],"tipo":"error","norma":"RG 5756/2025 Art. 3"},
        {"id":"EXPO-004","descripcion":"El carril de bodega compartida debe llamarse ATA CBC, no ATA CVC","patron":["ATA CVC"],"tipo":"advertencia","norma":"RG 5756/2025 Anexo §1.2"},
        {"id":"EXPO-005","descripcion":"Debe existir nodo de Ratificación de Autoría con token NF4","patron_ausente":["Ratificación de autoría","Ratificacion de autoria"],"tipo":"error","norma":"RG 5756/2025 §2.2"},
    ],
    "IMPO": [
        {"id":"IMPO-001","descripcion":"Transmisión IA del ATA MT debe incluir XFWB además del XFFM","patron":["XFFM"],"patron_ausente":["XFWB"],"tipo":"error","norma":"RG 3596/2014"},
        {"id":"IMPO-003","descripcion":"El timer de presentación automática debe ser 15 minutos desde confirmación de arribo","patron_ausente":["15 min","15min"],"tipo":"advertencia","norma":"RG 4517/2019 Art. 7"},
        {"id":"IMPO-005","descripcion":"Ratificación de Autoría debe requerir token NF4","patron_ausente":["NF4","nivel 4","token"],"tipo":"error","norma":"RG 4517/2019"},
    ]
}

SYSTEM_NORMATIVA = """Sos un experto en normativa aduanera argentina especializado en carga aérea.
Normativa clave: RG 3596/2014 (IA vía aérea, XFFM+XFWB, plazo 4hs), RG 4517/2019 (MANI SIM, generación automática, token NF4, 15min post-arribo), RG 5756/2025 (MANE exportación, registro post puesta a bordo, 3hs de partida, ATA CBC).
Respondé citando artículos. Si hay ambigüedad normativa, señalala explícitamente."""

SYSTEM_CORREOS = """Sos un asistente de redacción de correos institucionales para DI REPA de ARCA.
Contexto: proyecto VUA, circuito de carga aérea, XML IATA, MANI SIM y MANE.
Estilo: formal con externos/superiores; informal con colegas (primer nombre, "Abrazo" al cerrar).
Federico Cáceres firma como "Fede" en correos informales.
Incluí siempre: ASUNTO: [texto] al inicio."""

@vua_bp.route("/vua")
@login_required
@modulo_required("vua")
def vua_index():
    with get_db(HIST_DB, row_factory=True) as con:
        cronologia = [dict(r) for r in con.execute("SELECT * FROM vua_cronologia ORDER BY orden ASC, id ASC").fetchall()]
        ejes       = [dict(r) for r in con.execute("SELECT * FROM vua_ejes ORDER BY orden ASC").fetchall()]
        minutas    = [dict(r) for r in con.execute("SELECT id,fecha,asunto,lugar,creado_por,creado FROM vua_minutas ORDER BY creado DESC LIMIT 20").fetchall()]
    return render_template("vua.html", roles=ROLES_PREDEFINIDOS, cronologia=cronologia,
        ejes=ejes, minutas=minutas, campos_xfwb=CAMPOS_XFWB,
        role=session.get("role","admin"), username=session.get("username",""))

@vua_bp.route("/api/vua/minuta", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_minuta():
    data = request.json or {}
    asunto       = data.get("asunto","")
    fecha        = data.get("fecha", datetime.today().strftime("%d/%m/%Y"))
    lugar        = data.get("lugar","")
    participantes = data.get("participantes",[])
    temas        = data.get("temas",[])
    acuerdos     = data.get("acuerdos",[])
    proximos     = data.get("proximos",[])

    doc = generar_acta_word(
        "ACTA DE REUNIÓN", fecha, asunto, lugar, participantes,
        secciones=[("Temas tratados", temas), ("Acuerdos", acuerdos), ("Próximos pasos", proximos)],
        roles_predefinidos=ROLES_PREDEFINIDOS,
    )
    minuta_id = str(uuid.uuid4())[:8]
    fname = f"Acta_{fecha.replace('/','_')}_{asunto[:30].replace(' ','_')}_{minuta_id}.docx"
    ruta = os.path.join("/data/minutas", fname)

    # Mejora 5: intentar generar con Node (mejor formato); fallback a python-docx
    script = os.path.join(os.path.dirname(__file__), "generar_informe_vua.js")
    datos_minuta = {
        "fecha": fecha, "asunto": asunto, "lugar": lugar,
        "participantes": participantes, "temas": temas,
        "acuerdos": acuerdos, "proximos": proximos,
    }
    usó_node = False
    if os.path.exists(script):
        import tempfile as _tmp
        with _tmp.NamedTemporaryFile(suffix=".json", delete=False, mode="w", encoding="utf-8") as jf:
            json.dump(datos_minuta, jf, ensure_ascii=False)
            json_path = jf.name
        try:
            res = subprocess.run(["node", script, json_path, ruta, "minuta"],
                                 capture_output=True, text=True, encoding="utf-8",
                                 env={**os.environ, "LANG": "en_US.UTF-8", "NODE_OPTIONS": "--no-deprecation"},
                                 timeout=20)
            if res.returncode == 0 and os.path.exists(ruta):
                usó_node = True
        except Exception as e:
            logging.debug(f"Node.js generación Word falló (se usará método alternativo): {e}")
        finally:
            try: os.unlink(json_path)
            except: pass

    if not usó_node:
        # Fallback: python-docx básico (comportamiento anterior)
        doc.save(ruta)

    with get_db(HIST_DB) as con:
        con.execute("INSERT INTO vua_minutas VALUES (?,?,?,?,?,?,?,?,?,?,datetime('now'))",
            (minuta_id, fecha, asunto, lugar, json.dumps(participantes), json.dumps(temas),
             json.dumps(acuerdos), json.dumps(proximos), ruta, session.get("username","?")))

    # Construir sugerencia de cronología a partir de los datos de la reunión
    partic_nombres = ", ".join(
        p.get("nombre", p) if isinstance(p, dict) else str(p)
        for p in participantes
    )
    actividad_sugerida = asunto if asunto else "Reunión de trabajo VUA"
    crono_sugerida = {
        "fecha":        fecha,
        "actividad":    actividad_sugerida,
        "participantes": partic_nombres,
        "estado":       "Completado",
    }
    return jsonify({
        "ok":       True,
        "minuta_id": minuta_id,
        "download_url": f"/api/vua/minuta/{minuta_id}/download",
        "fname":    fname,
        "cronologia_sugerida": crono_sugerida,
    })

@vua_bp.route("/api/vua/correo", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_correo():
    data = request.json or {}
    instruccion = data.get("instruccion","")
    if not instruccion: return jsonify({"ok":False,"error":"Instrucción vacía"})
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=get_api_key(), http_client=httpx.Client(follow_redirects=True))
        msg = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1000,
            system=SYSTEM_CORREOS + contexto_repositorio("vua"), messages=[{"role":"user","content":instruccion}])
        return jsonify({"ok":True,"texto":msg.content[0].text})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@vua_bp.route("/api/vua/normativa", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_normativa():
    data = request.json or {}
    pregunta = data.get("pregunta","").strip()
    if not pregunta: return jsonify({"ok":False,"error":"Pregunta vacía"})
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=get_api_key(), http_client=httpx.Client(follow_redirects=True))
        msg = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1200,
            system=SYSTEM_NORMATIVA + contexto_repositorio("vua"), messages=[{"role":"user","content":pregunta}])
        return jsonify({"ok":True,"respuesta":msg.content[0].text})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@vua_bp.route("/api/vua/bpmn", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_bpmn():
    if "archivo" not in request.files: return jsonify({"ok":False,"error":"No se recibió archivo"})
    archivo = request.files["archivo"]
    circuito = request.form.get("circuito","AUTO")
    try:
        xml_content = archivo.read(6_000_000).decode("utf-8")
        if re.search(r'<!(DOCTYPE|ENTITY)', xml_content, re.IGNORECASE):
            return jsonify({"ok":False,"error":"XML no permitido: contiene DOCTYPE/ENTITY."})
        ET.fromstring(xml_content)
    except Exception as e:
        return jsonify({"ok":False,"error":f"XML inválido: {e}"})
    if circuito == "AUTO":
        upper = xml_content.upper()
        if "EXPORTACI" in upper and "MANE" in upper: circuito = "EXPO"
        elif "IMPORTACI" in upper and "MANI" in upper: circuito = "IMPO"
        else: return jsonify({"ok":False,"error":"No se pudo detectar el circuito."})
    reglas = REGLAS_BPMN.get(circuito,[])
    errores=[]; advertencias=[]
    for regla in reglas:
        hallado=False
        if "patron" in regla:
            for p in regla["patron"]:
                if p.lower() in xml_content.lower():
                    (errores if regla["tipo"]=="error" else advertencias).append(
                        {"id":regla["id"],"descripcion":regla["descripcion"],"norma":regla["norma"],"patron_encontrado":p})
                    hallado=True; break
        if "patron_ausente" in regla and not hallado:
            if not any(p.lower() in xml_content.lower() for p in regla["patron_ausente"]):
                (errores if regla["tipo"]=="error" else advertencias).append(
                    {"id":regla["id"],"descripcion":regla["descripcion"],"norma":regla["norma"],"patron_encontrado":"ausente"})
    return jsonify({"ok":True,"circuito":circuito,"errores":errores,"advertencias":advertencias,"total":len(errores)+len(advertencias)})

@vua_bp.route("/api/vua/xfwb")
@login_required
@modulo_required("vua")
def vua_xfwb():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Checklist XFWB"
    HDR=PatternFill("solid",fgColor="242D4F"); HDR_F=Font(bold=True,color="FFFFFF",size=10)
    bs=Side(style="thin",color="CCCCCC"); BORDER=Border(left=bs,right=bs,top=bs,bottom=bs)
    headers=["Tab","Campo","Campo XML","Norma","Obligatorio","Observación"]
    ws.append(headers)
    for ci,h in enumerate(headers,1):
        cell=ws.cell(1,ci); cell.fill=HDR; cell.font=HDR_F
        cell.alignment=Alignment(horizontal="center"); cell.border=BORDER
    ALT=PatternFill("solid",fgColor="EEF2F7")
    for ri,campo in enumerate(CAMPOS_XFWB,2):
        row=[campo["tab"],campo["campo"],campo["campo_xml"],campo["norma"],
             "Sí" if campo["obligatorio"] else "No",campo["observacion"]]
        for ci,val in enumerate(row,1):
            cell=ws.cell(ri,ci,val); cell.border=BORDER
            cell.fill=ALT if ri%2==0 else PatternFill(); cell.font=Font(size=10)
    for ci in range(1,len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width=25
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="Checklist_XFWB.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@vua_bp.route("/api/vua/informe")
@login_required
@modulo_required("vua")
def vua_informe():
    """Mejora 2: generación async del informe VUA — mismo patrón que SINTIA."""
    try:
        with get_db(HIST_DB, row_factory=True) as con:
            def _q(sql, fallback=[]):
                try: return [dict(r) for r in con.execute(sql).fetchall()]
                except: return fallback
            config     = _q("SELECT * FROM vua_config")
            ejes       = _q("SELECT * FROM vua_ejes ORDER BY orden ASC")
            equipo     = _q("SELECT * FROM vua_equipo WHERE activo=1 ORDER BY organismo, nombre ASC")
            cronologia = _q("SELECT * FROM vua_cronologia ORDER BY orden ASC, id ASC")
            glosario   = _q("SELECT * FROM vua_glosario ORDER BY termino ASC")
            riesgos    = _q("SELECT * FROM vua_riesgos WHERE activo=1 ORDER BY orden ASC")
            minutas    = _q("SELECT * FROM vua_minutas ORDER BY id DESC LIMIT 20")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error leyendo BD: {e}"}), 500

    datos = {"config": config, "ejes": ejes, "equipo": equipo,
             "cronologia": cronologia, "glosario": glosario,
             "riesgos": riesgos, "minutas": minutas}

    job_id = str(uuid.uuid4())[:8]
    job_create(job_id, "Generando informe VUA...", username=session.get("username", "?"))

    def _run_vua_informe(job_id, datos):
        log = job_status[job_id]["log"]
        try:
            with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w", encoding="utf-8") as jf:
                json.dump(datos, jf, ensure_ascii=False)
                json_path = jf.name
            out_path = json_path.replace(".json", ".docx")
            script   = os.path.join(os.path.dirname(__file__), "generar_informe_vua.js")
            result   = subprocess.run(["node", script, json_path, out_path],
                           capture_output=True, text=True, encoding="utf-8",
                           env={**os.environ, "LANG": "en_US.UTF-8", "NODE_OPTIONS": "--no-deprecation"},
                           timeout=60)
            if result.returncode != 0 or not os.path.exists(out_path):
                stderr_txt = result.stderr[:400] if result.stderr else "(sin stderr)"
                stdout_txt = result.stdout[:200] if result.stdout else "(sin stdout)"
                log.append(f"✗ Error Node (rc={result.returncode}): {stderr_txt} | stdout: {stdout_txt}")
                job_status[job_id]["status"] = "error"
                _job_persist(job_id)
                return
            fname = f"Informe_VUA_{datetime.today().strftime('%Y%m%d_%H%M')}_{job_id}.docx"
            os.makedirs(OUTPUT_FOLDER, exist_ok=True)
            dest = os.path.join(OUTPUT_FOLDER, fname)
            import shutil as _sh
            _sh.copy2(out_path, dest)
            try: os.unlink(out_path)
            except: pass
            job_status[job_id]["files"] = [dest]
            log.append(f"✓ Informe generado: {fname}")
            job_status[job_id]["status"] = "done"
            _job_persist(job_id)
            notificar_telegram(f"✓ Informe VUA listo ({job_status[job_id].get('username','?')})")
            try: os.unlink(json_path)
            except: pass
        except subprocess.TimeoutExpired:
            log.append("✗ Timeout generando informe")
            job_status[job_id]["status"] = "error"
            _job_persist(job_id)
            notificar_telegram(f"⚠️ Informe VUA falló por timeout ({job_status[job_id].get('username','?')})")
        except Exception as e:
            log.append(f"✗ {e}")
            job_status[job_id]["status"] = "error"
            _job_persist(job_id)
            notificar_telegram(f"⚠️ Informe VUA falló ({job_status[job_id].get('username','?')}): {e}")

    threading.Thread(target=_run_vua_informe, args=(job_id, datos)).start()
    return jsonify({"ok": True, "job_id": job_id})

@vua_bp.route("/api/vua/informe/download/<job_id>")
@login_required
@modulo_required("vua")
def vua_informe_download(job_id):
    """Descarga el informe VUA. Busca primero en memoria/SQLite, luego en disco por job_id."""
    # 1. Verificar estado del job (memoria del worker que lo creó, o SQLite si es otro worker)
    job = job_get(job_id)
    if job:
        if job["status"] != "done":
            return jsonify({"ok": False, "status": job["status"], "log": job.get("log",[])}), 202
        files = job.get("files", [])
        if files and os.path.exists(files[0]):
            fpath = files[0]
            fname = os.path.basename(fpath)
            if not fname.endswith(".docx"): fname = fname.rsplit(".", 1)[0] + ".docx"
            return send_file(fpath, as_attachment=True, download_name=fname,
                mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")

    # 2. Fallback: buscar en disco por job_id en el nombre del archivo
    #    (cubre reinicio del servidor entre generación y descarga)
    import glob
    if not re.match(r'^[a-zA-Z0-9]{1,32}$', job_id or ""):
        return jsonify({"ok": False, "error": "job_id inválido"}), 400
    patron = os.path.join(OUTPUT_FOLDER, f"*{job_id}*.docx")
    archivos = sorted(glob.glob(patron), key=os.path.getmtime, reverse=True)
    if archivos and os.path.exists(archivos[0]):
        return send_file(archivos[0], as_attachment=True,
            download_name=os.path.basename(archivos[0]),
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")

    # 3. Si no hay job en memoria y no hay archivo, dar error claro
    if not job:
        return jsonify({"ok": False,
            "error": "Sesión expirada o servidor reiniciado. Regenerá el informe."}), 404
    return jsonify({"ok": False, "error": "Archivo no encontrado en el servidor"}), 404


@vua_bp.route("/api/vua/cronologia", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_cronologia_get():
    with get_db(HIST_DB, row_factory=True) as con:
        rows=[dict(r) for r in con.execute(
            "SELECT *, CASE WHEN fecha GLOB '[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]' "
            "THEN substr(fecha,7,4)||substr(fecha,4,2)||substr(fecha,1,2) ELSE '00000000' END as _ord "
            "FROM vua_cronologia ORDER BY (estado='Pendiente') DESC, _ord DESC, id ASC").fetchall()]
    for r in rows: r.pop("_ord", None)
    return jsonify({"ok":True,"rows":rows})

@vua_bp.route("/api/vua/cronologia", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_cronologia_add():
    data=request.json or {}
    ok, err = validar_enum(data.get("estado"), ESTADOS_TAREA, "estado")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    fecha = _normalizar_fecha_a_ddmmaaaa(data.get("fecha","A definir")) if data.get("fecha") else "A definir"
    with get_db(HIST_DB) as con:
        cur=con.cursor()
        cur.execute("SELECT MAX(orden) FROM vua_cronologia")
        max_orden=cur.fetchone()[0] or 0
        cur.execute("INSERT INTO vua_cronologia (fecha,actividad,participantes,estado,orden,creado,modificado) VALUES (?,?,?,?,?,datetime('now'),datetime('now'))",
            (fecha,data.get("actividad",""),data.get("participantes",""),data.get("estado","Pendiente"),max_orden+1))
        new_id=cur.lastrowid
    return jsonify({"ok":True,"id":new_id})

@vua_bp.route("/api/vua/cronologia/<int:item_id>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_cronologia_update(item_id):
    data=request.json or {}
    if "fecha" in data and not _validar_fecha_ddmmaaaa(data["fecha"]):
        return jsonify({"ok":False,"error":f"Fecha inválida: '{data['fecha']}'. Formato requerido: dd/mm/aaaa."}), 400
    ok, err = validar_enum(data.get("estado"), ESTADOS_TAREA, "estado")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    campos = {k: v for k, v in data.items() if k in ("fecha","actividad","participantes","estado")}
    if not campos:
        return jsonify({"ok":False,"error":"Nada para actualizar"}), 400
    set_clause = ", ".join(f"{k}=?" for k in campos)
    with get_db(HIST_DB) as con:
        con.execute(f"UPDATE vua_cronologia SET {set_clause}, modificado=datetime('now') WHERE id=?",
            (*campos.values(), item_id))
    return jsonify({"ok":True})

@vua_bp.route("/api/vua/cronologia/<int:item_id>", methods=["DELETE"])
@login_required
@modulo_required("vua")
@admin_required()
def vua_cronologia_delete(item_id):
    with get_db(HIST_DB) as con:
        con.execute("DELETE FROM vua_cronologia WHERE id=?",(item_id,))
    return jsonify({"ok":True})

@vua_bp.route("/api/vua/minutas", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_minutas_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows=[dict(r) for r in con.execute("SELECT id,fecha,asunto,lugar,creado_por,creado FROM vua_minutas ORDER BY creado DESC").fetchall()]
    return jsonify({"ok":True,"rows":rows})

@vua_bp.route("/api/vua/minutas/<minuta_id>/download")
@login_required
@modulo_required("vua")
def vua_minuta_download(minuta_id):
    with get_db(HIST_DB, row_factory=True) as con:
        row=con.execute("SELECT * FROM vua_minutas WHERE id=?",(minuta_id,)).fetchone()
    if not row: return "No encontrada",404
    row=dict(row)
    if row.get("archivo") and os.path.exists(row["archivo"]):
        return send_file(row["archivo"],as_attachment=True,download_name=os.path.basename(row["archivo"]))
    return "Archivo no encontrado",404

@vua_bp.route("/api/vua/minutas/<minuta_id>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_minuta_delete(minuta_id):
    with get_db(HIST_DB) as con:
        row=con.execute("SELECT archivo FROM vua_minutas WHERE id=?",(minuta_id,)).fetchone()
        if row and row[0] and os.path.exists(row[0]):
            try: os.remove(row[0])
            except: pass
        con.execute("DELETE FROM vua_minutas WHERE id=?",(minuta_id,))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════
# RUTAS VUA BD DINAMICA
# ══════════════════════════════════════════════════════


# ── VUA Config ────────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/config", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_config_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute("SELECT * FROM vua_config ORDER BY clave").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/config/<clave>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_config_update(clave):
    data = request.json or {}
    contenido = data.get("contenido", "").strip()
    if not contenido: return jsonify({"ok": False, "error": "Contenido vacio"})
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_config SET contenido=?, modificado=datetime('now') WHERE clave=?", (contenido, clave))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/config/<clave>/mejorar", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_config_mejorar(clave):
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    with get_db(HIST_DB, row_factory=True) as con:
        row = con.execute("SELECT * FROM vua_config WHERE clave=?", (clave,)).fetchone()
    if not row: return jsonify({"ok": False, "error": "No encontrado"})
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))
        titulo = row["titulo"]
        contenido_actual = row["contenido"]
        # Mejora 3: system prompt con contexto institucional
        system_informe = (
            "Sos redactor de informes de gestión de proyectos para DI REPA de ARCA (Aduana Argentina). "
            "El proyecto es VUA (Ventanilla Única Aeroportuaria) para digitalización del circuito de carga aérea. "
            "Organismos involucrados: ARCA/DGA, VUCEA, SENASA, ORSNA, IATA, aerolíneas. "
            "Estilo: formal, preciso, en español rioplatense institucional. "
            "No agregues datos que no estén en el texto original. Solo mejorá la redacción y claridad."
        )
        prompt = (f"Mejorar la redacción de la sección '{titulo}' del informe de estado de situación del proyecto VUA. "
                  f"Conservá todos los datos y hechos del original. Devolvé solo el texto mejorado, sin encabezados ni explicaciones:\n\n{contenido_actual}")
        msg = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1500,
            system=system_informe + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        return jsonify({"ok": True, "texto": msg.content[0].text.strip()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ── VUA Ejes ──────────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/ejes", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_ejes_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute("SELECT * FROM vua_ejes ORDER BY orden").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/ejes/<eje_id>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_ejes_update_bd(eje_id):
    data = request.json or {}
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["nombre", "estado", "descripcion", "propuesta_vucea", "postura_aduana", "recomendacion"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(str(eje_id))
            con.execute("UPDATE vua_ejes SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/ejes", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_ejes_create():
    data = request.json or {}
    with get_db(HIST_DB, row_factory=True) as con:
        max_orden = con.execute("SELECT MAX(orden) FROM vua_ejes").fetchone()[0] or 0
        max_id = con.execute("SELECT id FROM vua_ejes ORDER BY orden DESC LIMIT 1").fetchone()
        try:
            last_num = float((max_id[0] if max_id else "0").replace(",","."))
            new_id = str(round(last_num + 0.1, 1))
        except: new_id = str(max_orden + 1)
        con.execute("INSERT INTO vua_ejes (id, nombre, estado, orden, descripcion, propuesta_vucea, postura_aduana, recomendacion) VALUES (?,?,?,?,?,?,?,?)",
            (new_id, data.get("nombre",""), data.get("estado","Pendiente"), max_orden + 1,
             data.get("descripcion",""), data.get("propuesta_vucea",""), data.get("postura_aduana",""), data.get("recomendacion","")))
    return jsonify({"ok": True, "id": new_id})

@vua_bp.route("/api/vua/ejes/<eje_id>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_ejes_delete(eje_id):
    with get_db(HIST_DB) as con:
        con.execute("DELETE FROM vua_ejes WHERE id=?", (str(eje_id),))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/ejes/<eje_id>/mejorar", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_eje_mejorar(eje_id):
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    with get_db(HIST_DB, row_factory=True) as con:
        eje = con.execute("SELECT * FROM vua_ejes WHERE id=?", (str(eje_id),)).fetchone()
    if not eje: return jsonify({"ok": False, "error": "No encontrado"})
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))
        nombre = eje["nombre"]; estado = eje["estado"]
        prompt = "Mejora nombre y estado del eje VUA. Solo JSON: {\"nombre\":\"...\",\"estado\":\"...\"}\nNOMBRE: " + nombre + "\nESTADO: " + estado
        msg = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=400,
            messages=[{"role": "user", "content": prompt}])
        resultado = json.loads(msg.content[0].text.strip().replace("```json","").replace("```",""))
        return jsonify({"ok": True, "nombre": resultado.get("nombre",""), "estado": resultado.get("estado","")})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ── VUA Equipo ────────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/equipo", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_equipo_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute(
            "SELECT * FROM vua_equipo WHERE activo=1 ORDER BY orden, organismo, nombre").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/equipo", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_equipo_create():
    data = request.json or {}
    with get_db(HIST_DB) as con:
        con.execute("INSERT INTO vua_equipo (nombre, cargo, organismo, email, activo) VALUES (?,?,?,?,1)",
            (data.get("nombre",""), data.get("cargo",""), data.get("organismo",""), data.get("email","")))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/equipo/<int:uid>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_equipo_update(uid):
    data = request.json or {}
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["nombre","cargo","organismo","email","activo"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(uid)
            con.execute("UPDATE vua_equipo SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/equipo/<int:uid>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_equipo_delete(uid):
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_equipo SET activo=0 WHERE id=?", (uid,))
    return jsonify({"ok": True})

# ── VUA Glosario ──────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/glosario", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_glosario_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute("SELECT * FROM vua_glosario ORDER BY orden, termino").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/glosario", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_glosario_create():
    data = request.json or {}
    with get_db(HIST_DB) as con:
        con.execute("INSERT INTO vua_glosario (termino, definicion, categoria) VALUES (?,?,?)",
            (data.get("termino",""), data.get("definicion",""), data.get("categoria","general")))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/glosario/<int:gid>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_glosario_update(gid):
    data = request.json or {}
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["termino","definicion","categoria"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(gid)
            con.execute("UPDATE vua_glosario SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/glosario/<int:gid>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_glosario_delete(gid):
    with get_db(HIST_DB) as con:
        con.execute("DELETE FROM vua_glosario WHERE id=?", (gid,))
    return jsonify({"ok": True})

# ── VUA Riesgos ───────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/riesgos", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_riesgos_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute(
            "SELECT * FROM vua_riesgos WHERE activo=1 ORDER BY orden").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/riesgos/<int:rid>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_riesgos_update(rid):
    data = request.json or {}
    ok, err = validar_enum(data.get("probabilidad"), NIVELES_PROBABILIDAD, "probabilidad")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    ok, err = validar_enum(data.get("impacto"), NIVELES_IMPACTO, "impacto")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["titulo","descripcion","mitigacion","probabilidad","impacto","activo"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(rid)
            con.execute("UPDATE vua_riesgos SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/riesgos", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_riesgos_create():
    data = request.json or {}
    ok, err = validar_enum(data.get("probabilidad"), NIVELES_PROBABILIDAD, "probabilidad")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    ok, err = validar_enum(data.get("impacto"), NIVELES_IMPACTO, "impacto")
    if not ok: return jsonify({"ok": False, "error": err}), 400
    with get_db(HIST_DB, row_factory=True) as con:
        max_orden = con.execute("SELECT MAX(orden) FROM vua_riesgos").fetchone()[0] or 0
        max_id = con.execute("SELECT MAX(id) FROM vua_riesgos").fetchone()[0] or 0
        codigo = f"R{max_id + 1:02d}"
        con.execute("INSERT INTO vua_riesgos (codigo, titulo, descripcion, mitigacion, probabilidad, impacto, activo, orden) VALUES (?,?,?,?,?,?,1,?)",
            (data.get("codigo", codigo), data.get("titulo",""), data.get("descripcion",""),
             data.get("mitigacion",""), data.get("probabilidad","Media"), data.get("impacto","Alto"), max_orden + 1))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/riesgos/<int:rid>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_riesgos_delete(rid):
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_riesgos SET activo=0 WHERE id=?", (rid,))
    return jsonify({"ok": True})

# ── VUA Correos rapidos ───────────────────────────────────────────────────────
@vua_bp.route("/api/vua/correos_rapidos", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_correos_rapidos_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute(
            "SELECT * FROM vua_correos_rapidos WHERE activo=1 ORDER BY orden").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/correos_rapidos", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_correos_rapidos_create():
    data = request.json or {}
    with get_db(HIST_DB) as con:
        con.execute("INSERT INTO vua_correos_rapidos (etiqueta, instruccion, activo) VALUES (?,?,1)",
            (data.get("etiqueta",""), data.get("instruccion","")))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/correos_rapidos/<int:cid>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_correos_rapidos_update(cid):
    data = request.json or {}
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["etiqueta","instruccion","activo"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(cid)
            con.execute("UPDATE vua_correos_rapidos SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/correos_rapidos/<int:cid>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_correos_rapidos_delete(cid):
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_correos_rapidos SET activo=0 WHERE id=?", (cid,))
    return jsonify({"ok": True})

# ── VUA Info ──────────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/info/<clave>", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_info_get(clave):
    with get_db(HIST_DB, row_factory=True) as con:
        row = con.execute("SELECT * FROM vua_info WHERE clave=?", (clave,)).fetchone()
    if not row: return jsonify({"ok": False, "error": "No encontrado"})
    return jsonify({"ok": True, "item": dict(row)})

@vua_bp.route("/api/vua/info/<clave>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_info_update(clave):
    data = request.json or {}
    contenido = data.get("contenido", "").strip()
    if not contenido: return jsonify({"ok": False, "error": "Contenido vacio"})
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_info SET contenido=?, modificado=datetime('now') WHERE clave=?", (contenido, clave))
    return jsonify({"ok": True})

# ── VUA Consultas frecuentes ──────────────────────────────────────────────────
@vua_bp.route("/api/vua/consultas_frecuentes", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_consultas_frecuentes_list():
    with get_db(HIST_DB, row_factory=True) as con:
        rows = [dict(r) for r in con.execute(
            "SELECT * FROM vua_consultas_frecuentes WHERE activo=1 ORDER BY orden").fetchall()]
    return jsonify({"ok": True, "rows": rows})

@vua_bp.route("/api/vua/consultas_frecuentes", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_consultas_frecuentes_create():
    """Permite guardar una consulta normativa ya resuelta como frecuente,
    igual que ya se puede hacer con correos_rapidos."""
    data = request.json or {}
    pregunta = (data.get("pregunta") or "").strip()
    if not pregunta:
        return jsonify({"ok": False, "error": "Falta la pregunta"})
    with get_db(HIST_DB) as con:
        con.execute("INSERT INTO vua_consultas_frecuentes (pregunta, respuesta, activo) VALUES (?,?,1)",
            (pregunta, data.get("respuesta","")))
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/consultas_frecuentes/<int:cid>", methods=["PUT"])
@login_required
@modulo_required("vua")
def vua_consultas_frecuentes_update(cid):
    data = request.json or {}
    with get_db(HIST_DB) as con:
        fields = []; params = []
        for f in ["pregunta","respuesta","activo"]:
            if f in data: fields.append(f + "=?"); params.append(data[f])
        if fields:
            params.append(cid)
            con.execute("UPDATE vua_consultas_frecuentes SET " + ", ".join(fields) + " WHERE id=?", params)
    return jsonify({"ok": True})

@vua_bp.route("/api/vua/consultas_frecuentes/<int:cid>", methods=["DELETE"])
@login_required
@modulo_required("vua")
def vua_consultas_frecuentes_delete(cid):
    with get_db(HIST_DB) as con:
        con.execute("UPDATE vua_consultas_frecuentes SET activo=0 WHERE id=?", (cid,))
    return jsonify({"ok": True})

# ── VUA Minuta IA ─────────────────────────────────────────────────────────────
@vua_bp.route("/api/vua/minuta_ia", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_minuta_ia():
    """Mejora 6: minuta_ia con contexto de minutas anteriores para detectar pendientes."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    data = request.json or {}
    asunto        = data.get("asunto", "")
    participantes = data.get("participantes", [])
    temas         = data.get("temas", [])
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))

        # Cargar últimas 5 minutas para contexto acumulado
        with get_db(HIST_DB, row_factory=True) as con:
            minutas_ant = [dict(r) for r in con.execute(
                "SELECT fecha, asunto, acuerdos, proximos FROM vua_minutas ORDER BY creado DESC LIMIT 5").fetchall()]

        ctx_minutas = ""
        if minutas_ant:
            ctx_minutas = "\n\nCONTEXTO — Últimas minutas del proyecto (para detectar pendientes y continuidad):\n"
            for m in reversed(minutas_ant):  # cronológico
                try:
                    acuerdos_prev = json.loads(m.get("acuerdos","[]") or "[]")
                    proximos_prev = json.loads(m.get("proximos","[]") or "[]")
                    ctx_minutas += (f"\n• {m['fecha']} — {m['asunto']}\n"
                                    f"  Acuerdos: {'; '.join(acuerdos_prev[:3])}\n"
                                    f"  Próximos pasos: {'; '.join(proximos_prev[:3])}\n")
                except: pass

        p_txt = "; ".join([p.get("nombre","") + " (" + p.get("cargo","") + ")" for p in participantes])
        t_txt = "\n".join(["- " + t for t in temas])

        prompt = (
            "Sos analista de DI REPA. Generá un borrador de acta para el proyecto VUA.\n"
            f"ASUNTO: {asunto}\n"
            f"PARTICIPANTES: {p_txt}\n"
            f"TEMAS TRATADOS HOY:\n{t_txt}"
            f"{ctx_minutas}\n\n"
            "Con ese contexto:\n"
            "1. Redactá los puntos tratados en esta reunión\n"
            "2. Identificá acuerdos concretos (con responsable si es posible)\n"
            "3. Definí próximos pasos, mencionando si alguno viene de reuniones anteriores y aún está pendiente\n\n"
            "Devolvé SOLO JSON válido (sin markdown):\n"
            "{\"temas_tratados\":[\"...\"]}\n"
            "{\"acuerdos\":[\"...\"]}\n"
            "{\"proximos_pasos\":[\"...\"]}\n"
            "{\"pendientes_anteriores\":[\"...\"]}\n"
            "Estilo: formal, español rioplatense institucional."
        )
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=1800,
            system="Sos un asistente experto en gestión de proyectos aduaneros para ARCA Argentina. Respondés solo con JSON válido, sin texto adicional." + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        texto = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
        resultado = json.loads(texto)
        return jsonify({"ok": True, **resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})



# ── Importar minuta desde Word con IA ────────────────────────────────────────
@vua_bp.route("/api/vua/minuta/importar", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_minuta_importar():
    """Recibe un .docx, extrae el texto y usa la IA para estructurarlo en campos de minuta."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    if "archivo" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió archivo"})

    archivo = request.files["archivo"]
    if not archivo.filename.endswith(".docx"):
        return jsonify({"ok": False, "error": "Solo se aceptan archivos .docx"})

    # Extraer texto del Word con python-docx
    try:
        from docx import Document as DocxDoc
        import io as _io
        doc_bytes = archivo.read()
        docx_doc  = DocxDoc(_io.BytesIO(doc_bytes))
        # Extraer párrafos y tablas
        partes = []
        for p in docx_doc.paragraphs:
            txt = p.text.strip()
            if txt: partes.append(txt)
        for tabla in docx_doc.tables:
            for fila in tabla.rows:
                celda_txt = " | ".join(c.text.strip() for c in fila.cells if c.text.strip())
                if celda_txt: partes.append(celda_txt)
        texto_completo = "\n".join(partes)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error leyendo el Word: {e}"})

    if not texto_completo.strip():
        return jsonify({"ok": False, "error": "El documento está vacío o no se pudo extraer texto"})

    # Llamar a la IA para estructurar
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))

        prompt = (
            "El siguiente texto es una minuta/acta de reunión del proyecto VUA (Ventanilla Única Aeroportuaria — ARCA Argentina).\n"
            "Extraé y estructurá la información en el JSON solicitado.\n\n"
            f"TEXTO DE LA MINUTA:\n{texto_completo[:6000]}\n\n"
            "Devolvé SOLO este JSON válido (sin markdown ni texto adicional):\n"
            "{\n"
            '  "asunto": "título o asunto principal de la reunión",\n'
            '  "fecha": "fecha en formato YYYY-MM-DD si la encontrás, sino vacío",\n'
            '  "lugar": "lugar o modalidad (ej: Videoconferencia, Sala 3 Paseo Colón)",\n'
            '  "participantes": [\n'
            '    {"nombre": "Nombre completo o sigla del organismo", "cargo": "cargo si está disponible"}\n'
            '  ],\n'
            '  "temas_tratados": ["tema 1", "tema 2"],\n'
            '  "acuerdos": ["acuerdo 1", "acuerdo 2"],\n'
            '  "proximos_pasos": ["paso 1", "paso 2"]\n'
            "}\n\n"
            "Reglas:\n"
            "- temas_tratados: los temas principales que se discutieron, uno por ítem, en forma concisa\n"
            "- acuerdos: compromisos concretos que se tomaron en la reunión\n"
            "- proximos_pasos: tareas o acciones pendientes mencionadas\n"
            "- Si la fecha está escrita en texto (ej: '11 de junio de 2026'), convertila a YYYY-MM-DD\n"
            "- Español rioplatense, con tildes y caracteres especiales correctos (á, é, í, ó, ú, ñ), sin markdown"
        )

        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=2000,
            system="Sos un asistente experto en gestión de proyectos aduaneros para ARCA Argentina. Extraés información estructurada de minutas institucionales. Respondés solo con JSON válido." + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}]
        )
        texto_resp = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
        resultado  = json.loads(texto_resp)
        return jsonify({"ok": True, **resultado})

    except json.JSONDecodeError as e:
        return jsonify({"ok": False, "error": f"La IA no devolvió JSON válido: {e}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── VUA Resumen ejecutivo generado por IA (Mejora 7) ─────────────────────────
@vua_bp.route("/api/vua/config/resumen_ejecutivo/generar", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_resumen_generar():
    """Mejora 7: genera el resumen ejecutivo automáticamente desde el estado actual del proyecto."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    try:
        import anthropic, httpx
        with get_db(HIST_DB, row_factory=True) as con:
            ejes       = [dict(r) for r in con.execute("SELECT id, nombre, estado FROM vua_ejes ORDER BY orden").fetchall()]
            riesgos    = [dict(r) for r in con.execute("SELECT titulo, probabilidad, impacto FROM vua_riesgos WHERE activo=1").fetchall()]
            cronologia = [dict(r) for r in con.execute(
                "SELECT fecha, actividad, estado FROM vua_cronologia ORDER BY orden DESC LIMIT 10").fetchall()]
            minutas    = [dict(r) for r in con.execute(
                "SELECT fecha, asunto, proximos FROM vua_minutas ORDER BY creado DESC LIMIT 3").fetchall()]

        ult_actividad = next((c for c in cronologia if c["estado"].lower() == "completado"), {})
        prox_actividad = next((c for c in reversed(cronologia) if c["estado"].lower() == "pendiente"), {})

        ejes_txt = "\n".join([f"• {e['id']} {e['nombre']}: {e['estado']}" for e in ejes])
        riesgos_txt = "\n".join([f"• {r['titulo']} (Prob: {r['probabilidad']}, Imp: {r['impacto']})" for r in riesgos[:5]])
        
        pendientes_minuta = []
        for m in minutas:
            try:
                proximos = json.loads(m.get("proximos","[]") or "[]")
                pendientes_minuta.extend(proximos[:2])
            except: pass

        prompt = (
            "Redactá el Resumen Ejecutivo del informe de estado de situación del proyecto VUA "
            "(Ventanilla Única Aeroportuaria — ARCA/Aduana Argentina) para un informe formal de gestión.\n\n"
            f"EJES DEL PROYECTO:\n{ejes_txt}\n\n"
            f"RIESGOS ACTIVOS:\n{riesgos_txt}\n\n"
            f"ÚLTIMA ACTIVIDAD COMPLETADA: {ult_actividad.get('fecha','')} — {ult_actividad.get('actividad','')}\n"
            f"PRÓXIMA ACTIVIDAD PROGRAMADA: {prox_actividad.get('fecha','')} — {prox_actividad.get('actividad','')}\n\n"
            f"COMPROMISOS PENDIENTES DE MINUTAS: {'; '.join(pendientes_minuta[:5])}\n\n"
            "El resumen debe: describir el estado general del proyecto, mencionar los ejes más avanzados y los pendientes, "
            "señalar los principales riesgos y los próximos hitos. "
            "Extensión: 3-4 párrafos. Estilo: formal, español rioplatense institucional. "
            "No uses bullet points — solo prosa. No incluyas títulos."
        )
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=1200,
            system="Sos redactor de informes institucionales de ARCA Argentina. Redactás en prosa formal, español rioplatense, sin markdown." + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        return jsonify({"ok": True, "texto": msg.content[0].text.strip()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── VUA Validador BPMN con IA (Mejora 8) ─────────────────────────────────────
@vua_bp.route("/api/vua/bpmn/ia", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_bpmn_ia():
    """Mejora 8: validación BPMN profunda con IA — complementa el validador de regex."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    if "archivo" not in request.files: return jsonify({"ok": False, "error": "No se recibió archivo"})
    archivo  = request.files["archivo"]
    circuito = request.form.get("circuito", "AUTO")
    try:
        xml_content = archivo.read().decode("utf-8")
        import xml.etree.ElementTree as ET
        ET.fromstring(xml_content)
    except Exception as e:
        return jsonify({"ok": False, "error": f"XML inválido: {e}"})

    if circuito == "AUTO":
        upper = xml_content.upper()
        if   "EXPORTACI" in upper and "MANE" in upper: circuito = "EXPO"
        elif "IMPORTACI" in upper and "MANI" in upper: circuito = "IMPO"
        else: return jsonify({"ok": False, "error": "No se pudo detectar el circuito automáticamente."})

    # Primero correr el validador rápido de regex
    reglas = REGLAS_BPMN.get(circuito, [])
    errores_regex = []; advertencias_regex = []
    for regla in reglas:
        hallado = False
        if "patron" in regla:
            for pat in regla["patron"]:
                if pat.lower() in xml_content.lower():
                    (errores_regex if regla["tipo"]=="error" else advertencias_regex).append(
                        {"id":regla["id"],"descripcion":regla["descripcion"],"norma":regla["norma"],"fuente":"regex"})
                    hallado = True; break
        if "patron_ausente" in regla and not hallado:
            if not any(p.lower() in xml_content.lower() for p in regla["patron_ausente"]):
                (errores_regex if regla["tipo"]=="error" else advertencias_regex).append(
                    {"id":regla["id"],"descripcion":regla["descripcion"],"norma":regla["norma"],"fuente":"regex"})

    # Luego análisis profundo con IA
    try:
        import anthropic, httpx
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))

        # Truncar XML a 8000 chars para no superar contexto
        xml_truncado = xml_content[:8000] + ("\n[...truncado...]" if len(xml_content) > 8000 else "")

        prompt = (
            f"Analizá este diagrama BPMN del circuito {circuito} de carga aérea (proyecto VUA, ARCA Argentina).\n\n"
            f"XML DEL BPMN:\n{xml_truncado}\n\n"
            "Verificá específicamente:\n"
            "1. Asignación correcta de tareas a swimlanes (ATA MT, ATA CBC, ATA CVC, ATA AGT, ADUANA, SENASA)\n"
            "2. Nombres de nodos según nomenclatura aduanera argentina (MANE, MANI SIM, XFFM, XFWB, PATAI, OFTAI)\n"
            "3. Completitud del flujo: inicio → transmisión anticipada → arribo → validación → despacho\n"
            "4. Presencia de elementos obligatorios según la normativa (RG 3596/2014, RG 4517/2019, RG 5756/2025)\n"
            "5. Inconsistencias lógicas en el flujo (decisiones sin todas sus ramas, tareas sin conexión)\n\n"
            "Devolvé SOLO JSON válido:\n"
            "{\"errores_ia\":[{\"id\":\"IA-001\",\"descripcion\":\"...\",\"norma\":\"...\",\"sugerencia\":\"...\"}],\"advertencias_ia\":[...],\"observaciones\":\"...resumen general...\"}"
        )
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=2000,
            system=SYSTEM_NORMATIVA + "\nRespondés solo con JSON válido, sin texto adicional." + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        texto = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
        resultado_ia = json.loads(texto)
    except Exception as e:
        resultado_ia = {"errores_ia": [], "advertencias_ia": [], "observaciones": f"Análisis IA no disponible: {e}"}

    return jsonify({
        "ok": True,
        "circuito": circuito,
        "errores":       errores_regex + resultado_ia.get("errores_ia", []),
        "advertencias":  advertencias_regex + resultado_ia.get("advertencias_ia", []),
        "observaciones": resultado_ia.get("observaciones",""),
        "total": len(errores_regex) + len(advertencias_regex) + len(resultado_ia.get("errores_ia",[])) + len(resultado_ia.get("advertencias_ia",[])),
    })


# ── VUA Acuerdos pendientes (Mejora 9) ────────────────────────────────────────
@vua_bp.route("/api/vua/acuerdos/pendientes", methods=["GET"])
@login_required
@modulo_required("vua")
def vua_acuerdos_pendientes():
    """Mejora 9: detecta compromisos sin seguimiento cruzando minutas con IA."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    try:
        import anthropic, httpx
        with get_db(HIST_DB, row_factory=True) as con:
            minutas = [dict(r) for r in con.execute(
                "SELECT fecha, asunto, acuerdos, proximos FROM vua_minutas ORDER BY creado ASC").fetchall()]

        if len(minutas) < 2:
            return jsonify({"ok": True, "pendientes": [], "mensaje": "Se necesitan al menos 2 minutas para detectar pendientes."})

        # Construir resumen de todas las minutas
        minutas_txt = ""
        for m in minutas:
            try:
                acuerdos = json.loads(m.get("acuerdos","[]") or "[]")
                proximos = json.loads(m.get("proximos","[]") or "[]")
                minutas_txt += f"\n--- {m['fecha']} — {m['asunto']} ---\n"
                if acuerdos: minutas_txt += "Acuerdos: " + " | ".join(acuerdos) + "\n"
                if proximos: minutas_txt += "Próximos pasos: " + " | ".join(proximos) + "\n"
            except: pass

        prompt = (
            "Analizá estas minutas del proyecto VUA (en orden cronológico) y detectá compromisos o próximos pasos "
            "que aparecen en reuniones anteriores pero NO tienen evidencia de resolución en reuniones posteriores.\n\n"
            f"MINUTAS:\n{minutas_txt}\n\n"
            "Devolvé SOLO JSON válido:\n"
            "{\"pendientes\":[{\"descripcion\":\"...\",\"origen\":\"fecha — reunión\",\"estado\":\"Sin evidencia de cierre\",\"prioridad\":\"Alta|Media|Baja\"}],"
            "\"resueltos_recientes\":[{\"descripcion\":\"...\",\"cerrado_en\":\"fecha\"}]}"
        )
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=2000,
            system="Sos analista de seguimiento de proyectos. Identificás compromisos y verificás su cumplimiento. Respondés solo con JSON válido." + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        texto = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
        resultado = json.loads(texto)
        return jsonify({"ok": True, **resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── VUA Sugerencia de mitigación de riesgos (Mejora 10) ──────────────────────
@vua_bp.route("/api/vua/riesgos/<int:rid>/mitigacion_ia", methods=["POST"])
@login_required
@modulo_required("vua")
def vua_riesgo_mitigacion_ia(rid):
    """Mejora 10: sugiere estrategias de mitigación específicas para un riesgo dado el contexto del proyecto."""
    api_key = get_api_key()
    if not api_key: return jsonify({"ok": False, "error": "API key no configurada"})
    try:
        import anthropic, httpx
        with get_db(HIST_DB, row_factory=True) as con:
            riesgo = con.execute("SELECT * FROM vua_riesgos WHERE id=?", (rid,)).fetchone()
            ejes   = [dict(r) for r in con.execute("SELECT nombre, estado FROM vua_ejes ORDER BY orden").fetchall()]
        if not riesgo: return jsonify({"ok": False, "error": "Riesgo no encontrado"})
        riesgo = dict(riesgo)

        ejes_txt   = " | ".join([f"{e['nombre']} ({e['estado']})" for e in ejes])
        prompt = (
            "Proyecto VUA — Ventanilla Única Aeroportuaria (ARCA Argentina, carga aérea internacional).\n"
            f"Ejes del proyecto: {ejes_txt}\n\n"
            "RIESGO A MITIGAR:\n"
            f"Título: {riesgo.get('titulo','')}\n"
            f"Descripción: {riesgo.get('descripcion','')}\n"
            f"Probabilidad: {riesgo.get('probabilidad','')} | Impacto: {riesgo.get('impacto','')}\n"
            f"Mitigación actual: {riesgo.get('mitigacion','')}\n\n"
            "Sugerí 3 estrategias de mitigación concretas y específicas para el contexto aduanero argentino. "
            "Para cada una indicá: acción concreta, responsable sugerido (organismo), y plazo estimado. "
            "Devolvé solo el texto de las 3 estrategias en prosa, numeradas, sin JSON."
        )
        client = anthropic.Anthropic(api_key=api_key, http_client=httpx.Client(follow_redirects=True))
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=800,
            system=SYSTEM_NORMATIVA + contexto_repositorio("vua"),
            messages=[{"role": "user", "content": prompt}])
        return jsonify({"ok": True, "sugerencias": msg.content[0].text.strip()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
